# print("I am Zanga")
# print('o----')
# print(' ||||')
# print('*' * 10)
#
# patient_fullname = 'Silicon'
# patient_age = 20
# is_new = True
from operator import truediv

# name = input('What is your name? ')
# colour = input('what is your favourite color? ')
# print(name + ' likes ' + colour)

# birth_year = input('Birth Year: ')
# print(type(birth_year))
# age = 2019 - int(birth_year)
# print(type(age))
# print(age)

# user_weight = input('Weight (lbs): ')
# kilogram = int(user_weight) * 0.4536
# print(kilogram)

# course = '''
# Hi John,
# Here's a testing email
# Thanks
# '''
# print(course)

# course = 'Testing index numbers'
# print(course[0])
# print(course[-1])
# print(course[0:3)
# name = 'Jennifer'
# print(name[1:-1])

# course = 'Python for Beginners'
# another = course[:]
# print(another)


# formatted strings
# first = 'John'
# last = 'Smith'
# message = f'{first} {last} is a coder'
# print(message)


# String Length
# name = 'Odusote is a man'
# print(len(name)) count number of characters
# print(name.upper()) uppercase
# print(name.lower()) lowercase
# print(name.replace('is a man', 'is a Professor'))
# print(name.replace('Odusote', 'Adekunle'))


# Boolean
# name = 'Oreoluwa is a girl'
# print('girl' in name)
# print('boy' in name)


# Arithmetic Operations
# print(10 * 3)
# x = 10 + 3 * 2 ** 2
# print(x)
# exponentiation 2 ** 3
# multiplication or division
# addition or subtraction


# Math Function
# x = 2.9
# print(round(x))
# print(abs(-2.9)) // Always returns a positive number


# If Statements
# is_hot = False
# is_cold = False
#
# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("It's a cold day")
#     print("Wear warm clothes")
# else:
#     print("It's a Lovely day")
#
# print('Enjoy your day')


# Exercise
# good_credit = True
# bad_credit = False
#
# if good_credit:
#     print("Pay 10%")
# elif bad_credit:
#     print("Pay 20%")
# else:
#     print("Pay the down payment")

# # Solution
# price = 1000000
# has_good_credit = False
#
# if has_good_credit:
#     down_payment = 0.1 * price
# else:
#     down_payment = 0.2 * price
# print(f"Down Payment is: ${down_payment}")


# # has_high_income = False
# has_good_credit = True
# has_criminal_record = False
#
# # if has_high_income and has_good_credit:
# #     print("Eligible for Loan")
#
# # if has_high_income or has_good_credit:
# #     print("Eligible for Loan")
#
# if has_good_credit and not has_criminal_record:
#     print("Eligible for Loan")
#
# # AND: both condition should be true
# # OR: at least one should be true
# # NOT: one condition should be false

# name = "Rid"
#
# if len(name) < 3:
#     print("name must be at least 3 characters")
# elif len(name) > 50:
#     print("name can be a maximum of 50 characters")
# else:
#     print("name looks good!")



# # while loop
# i = 1
# while i <= 5:
#     print(i)
#     i = i + 1
# print("Done")
#
# i = 1
# while i <= 5:
#     print('*' * i)
#     i = i + 1
# print("Done")



# # For Loop
# for item in 'Python':
#     print(item)
#
# for name in ['Mack, Rid, John, Bims']:
#     print(name)
#
# for item in range(10):
#     print(item)
#
#
# prices = [10, 20, 30]
# total = 0
# for price in prices:
#     total += price
# print(f"Total: {total}")


# # # Nested Loops
# for x in range(4):
#     for y in range(3):
#         print(f"({x}, {y})")

#
# numbers = [5, 2, 5, 2, 2]
# for x_count in numbers:
#     print('x' * x_count)
#
# numbers = [5, 2, 5, 2, 2]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)


# ## List
# names = ['John', 'Ridwan', 'Dare', 'Seun']
# names[0] = 'Jon'
# print(names[1:2])
# print(names[2])
# print(names)
#
# numbers = [3, 6, 2, 8, 4, 20]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)


## Dimensional List (2D)
# matrix = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9]
# ]
# # matrix[0][1] = 20
# # print(matrix[0][1])
# for row in matrix:
#     for item in row:
#         print(item)


# ## List Methods or List Functions
# numbers = [5, 2, 2, 4, 1, 7, 4]
# # numbers.append(20)
# # numbers.insert(0, 10)
# # numbers.remove(7)
# # numbers.pop() #removes last number
# # print(numbers)
#
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# # Tuples (they are immutable)
# numbers = (1, 2, 3)
# print(numbers[0])

# # Unpacking
# coordinates = (1, 2, 3)
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
#
# # or
# x, y, z = coordinates
# print(x)


# # Dictionaries
# customer = {
#     "name": "John Smith",
#     "age": 30,
#     "is_verified": True
# }
# customer["name"] = "Ridwan"
# print(customer["name"])
# print(customer["age"])
# print(customer.get("birthdate", "Jan 1 1980"))
#
# # Exercise
# phone = input("Phone: ")
# digits_mapping = {
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four"
# }
# output = ""
# for ch in phone:
#     output += digits_mapping.get(ch, "!") + " "
# print(output)


# Emoji Converter
# message = input("> ")
# words = message.split(' ')
# emojis = {
#     ":)": "😃",
#     ":(": "😔",
#     ":!": "😡"
# }
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)


# # Functions
# def greet_user():
#     print("Hi there!")
#     print("Welcome aboard.")
#
# print("Start")
# greet_user()
# print("Finish")


# # Parameter
# def greet_user(first_name, last_name):
#     print(f"Hi {first_name} {last_name}!")
#     print("Welcome aboard.")
#
# print("Start")
# greet_user("Ridwan", "Akins")
# print("Finish")


# Return Statements
# def square(number):
#     return number * number
#
# print(square(3))


# Exercise

# def emoji_converter(message):
#     words = message.split(' ')
#     emojis = {
#         ":)": "😃",
#         ":(": "😔",
#         ":!": "😡"
#     }
#     output = ""
#     for word in words:
#         output += emojis.get(word, word) + " "
#     return output
#
# message = input("> ")
# print(emoji_converter(message))



# # Exceptions
# try:
#     age = int(input('Age: '))
#     income = 20000
#     risk = income / age
#     print(risk)
# except ZeroDivisionError:
#     print('Age cannot be 0.')
# except ValueError:
#     print('Invalid Value')


# Classes
# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print(move)
#
#     def draw(self):
#         print(draw)
#
# point = Point(10, 20)
# point.x = 50
# print(point.x)

# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.x)
#
# point2 = Point()
# point2.x = 1
# print(point2.x)


# Exercise
# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"Hi, I am {self.name}")
#
# john = Person("John Doe")
# john.talk()
#
# bob = Person("Bob Smith")
# bob.talk()


# Inheritance
# class Mammal:
#     def walk(self):
#         print("walk")
#
# class Dog(Mammal):
#     def bark(self):
#         print("bark")
#
# class Cat(Mammal):
#     def meow(self):
#         print("meow")
#
# cat1 = Cat()
# cat1.walk()
# cat1.meow()
#
# dog1 = Dog()
# dog1.walk()
# dog1.bark()

#
# # Modules in Python from converter.py
# def lbs_to_kg(weight):
#     return weight * 0.45
#
# def kg_to_lbs(weight):
#     return weight / 0.45



# Exercise importing from utils.py
# from utils import find_max
#
# numbers = [10, 3, 6, 2]
# maximum = find_max(numbers)
# print(maximum(numbers))



# # Exercise module from ecommerce/shipping
# import ecommerce.shipping
# # or
# from ecommerce.shipping import calc_shipping
# # or
# from ecommerce import shipping
#
#
# ecommerce.shipping.calc_shipping()
# # or
# calc_shipping()
# # oor
# shipping.calc_shipping()


# Working with RANDOM module
# import random
# for i in range(3):
#     print(random.randint(10, 20))
#
# # Picking Randomly
# members = ['John', 'Mary', 'Bob', 'Josh']
# leader = random.choice(members)
# print(leader)
#
# # Rolling Dice
# class Dice:
#     def roll(self):
#         first = random.randint(1,6)
#         second = random.randint(1, 6)
#         return first, second
#
# dice = Dice()
# print(dice.roll())


# Pathlib
# Absolute path (/usr/local/bin)
# Relative path (file)
#
# from pathlib import Path
#
# path = Path()
# for file in path.glob('*.py'):
#     print(file)




# Excel Transaction Project (Example 1)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)

# for row in range(2, sheet.max_row + 1):
#     print(row)

# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     print(cell.value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')

# print(sheet.max_row)
# print(cell.value)





# Excel Transaction Project (Example 2)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)

    # for row in range(2, sheet.max_row + 1):
    #     print(row)

    # for row in range(2, sheet.max_row + 1):
    #     cell = sheet.cell(row, 3)
    #     print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

    # print(sheet.max_row)
    # print(cell.value)
